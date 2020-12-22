# coding=utf-8
# @Time : 2020/12/22 18:23
# @Author : zyk
# @Email : zhangyongke1105@163.com
# @File : file_read.py
# @Software: PyCharm
import openpyxl


def user_info():
    """用户信息文件，把用户信息保存在文件中，从文件中读取需要登录的用户名和密码"""
    # 使用字段存储用户信息
    user_information = {}
    # 打开文件
    try:
        wb = openpyxl.load_workbook(r"F:\\workplace\\GitWork\\python_project\\test_data\\工作表.xlsx")
    except FileNotFoundError as e:
        print(e)
    else:
        """使用wb.sheetnames获取所有sheet页，避免警告all_sheet_names = wb.sheetnames;
        使用wb[sheetname]获取对应的sheet页，可以避免警告
        # print(all_sheet_names)
        """
        sheet = wb["username"]

        # 获取当前sheet页，共有多少行和列
        rows = sheet.max_row
        # cows = sheet.max_column
        # print(row, cow)

        # 定位单元格cell,根据行列读取测试数据，去掉title行，从第2行开始。
        for i in range(2, rows+1, 1):
            username = sheet.cell(i, 1).value
            password = sheet.cell(i, 2).value
            user_information[username] = password
        # print(user_information)
        return user_information


def enterprise_read():
    batches = []   # 批次字段，把表格返回的数据，存放在字典中
    enterprise_names = []  # 企业名称，把表格返回的数据，存放在字典中
    credit_codes = []  # 企业名称，把表格返回的数据，存放在字典中
    wb = openpyxl.load_workbook(r"F:\\workplace\\jichupingtai\\file\\查询条件.xlsx")
    # all_sheet_names = wb.sheetnames
    # print(all_sheet_names)
    # 打开sheet页
    sheet = wb["企业白名单查询"]
    # 获取打开sheet页的行数和列数
    rows = sheet.max_row
    # cows = sheet.max_column
    # 定位单元格，获取单元格数据
    # for i in range(2, rows+1, 1):
    for i in range(2, 5, 1):
        a = sheet.cell(i, 4).value
        # b = sheet.cell(i, 2).value
        # print(a)
        batches.append(a)
    for j in range(5, 8, 1):
        b = sheet.cell(j, 4).value
        enterprise_names.append(b)
    for k in range(8, rows+1, 1):
        c = sheet.cell(k, 4).value
        credit_codes.append(c)
    # print(batches)
    return batches, enterprise_names, credit_codes


# user_info()
# data = enterprise_read()
# print(data)
