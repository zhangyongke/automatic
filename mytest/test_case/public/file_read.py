# coding=utf-8
# @Time : 2020/12/22 18:23
# @Author : zyk
# @Email : zhangyongke1105@163.com
# @File : file_read.py
# @Software: PyCharm
import openpyxl


def user_info(row):
    """用户信息文件，把用户信息保存在文件中，从文件中读取需要登录的用户名和密码"""
    # 使用字段存储用户信息
    # user_information = {}
    # 打开文件
    try:
        wb = openpyxl.load_workbook(r"F:\\workplace\\GitWork\\python_project\\test_data\\工作表.xlsx")
    except FileNotFoundError as e:
        print(e)
    else:
        """使用wb.sheetnames获取所有sheet页，避免警告;使用wb[sheetname]获取对应的sheet页，可以避免警告
        # print(all_sheet_names)
        """
        sheet = wb["user_information"]

        # 获取当前sheet页，共有多少行和列
        # rows = sheet.max_row
        # cows = sheet.max_column
        # print(row, cow)
        user_information = sheet.cell(row, 3).value
        username, password = user_information.split("，")
        if username == "null":
            username = ""
        if password == "null":
            password = ""

        # user_information[username] = password
        # # 定位单元格cell,根据行列读取测试数据，去掉title行，从第2行开始。
        # for i in range(4, rows+1, 1):
        #     user_information = sheet.cell(i, 3).value
        #     username, password = user_information.split("，")
        #     # username = sheet.cell(i, 1).value
        #     # password = sheet.cell(i, 2).value
        #     user_information[username] = password
        # return user_information
        return username, password


def open_fie(filepath, sheet_name):
    """打开文件操作"""
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    rows = sheet.max_row
    cows = sheet.max_column
    return sheet


# a = user_info(4)
# print(a)
# print(type(a))

